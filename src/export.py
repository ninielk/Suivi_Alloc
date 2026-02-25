"""
export.py — Export du tableau Alloc recalculé vers un fichier .xlsx formaté.
"""

from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.calculs import ROW_DEFS

# ──────────────────────────────────────────────────────────────────────────────
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILLS = {
    "blue":   PatternFill(fgColor="2E75B6", fill_type="solid"),
    "white":  PatternFill(fgColor="FFFFFF", fill_type="solid"),
    "total":  PatternFill(fgColor="1F3864", fill_type="solid"),
    "header": PatternFill(fgColor="1F3864", fill_type="solid"),
}
FONTS = {
    "blue":   Font(bold=True, color="FFFFFF", name="Calibri", size=11),
    "white":  Font(color="000000",            name="Calibri", size=11),
    "total":  Font(bold=True, color="FFFFFF", name="Calibri", size=11),
    "header": Font(bold=True, color="FFFFFF", name="Calibri", size=11),
}

HEADERS = [
    "Catégorie d'investissement",
    "Allocation cible (%)",
    "Marge de manœuvre",
    "Retraitement (M€)",
    "Allocation (M€)",
    "Allocation (%)",
    "VNC (M€)",
    "VNC (%)",
]

COL_WIDTHS = {"A": 42, "B": 18, "C": 20, "D": 18, "E": 18, "F": 14, "G": 14, "H": 12}


def _cell(ws, row, col, value, row_type, h_align="right", num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill   = FILLS[row_type]
    c.font   = FONTS[row_type]
    c.border = BORDER
    c.alignment = Alignment(horizontal=h_align, vertical="center")
    if num_fmt:
        c.number_format = num_fmt
    return c


def export_to_excel(results: dict) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alloc"

    # ── En-têtes ─────────────────────────────────────────────────────────────
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill      = FILLS["header"]
        c.font      = FONTS["header"]
        c.border    = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    # ── Lignes de données ─────────────────────────────────────────────────────
    for i, (row_num, label, row_type, detail) in enumerate(ROW_DEFS):
        r = results.get(row_num)
        if not r:
            continue
        erow = i + 2

        e_val = r["E"]
        # Normalise le % si stocké en décimal
        if isinstance(e_val, (int, float)) and e_val and abs(e_val) <= 1:
            e_num_fmt = "0%"
        else:
            e_num_fmt = "0%"

        d_val = r["D"] if (row_type == "white" and detail == "normal") else None

        row_data = [
            (label,  "left",  None),
            (e_val,  "right", e_num_fmt),
            (r["F"], "right", "@"),           # Marge de manoeuvre — texte brut
            (d_val,  "right", '#,##0'),
            (r["G"], "right", '#,##0'),
            (r["H"], "right", '0.0%'),
            (r["J"] if r["J"] is not None else None, "right", '#,##0'),
            (r["K"] if r["K"] is not None else None, "right", '0.0%'),
        ]

        for col, (val, align, fmt) in enumerate(row_data, 1):
            _cell(ws, erow, col, val, row_type, h_align=align, num_fmt=fmt)

    # ── Dimensions colonnes ───────────────────────────────────────────────────
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ── Freeze pane + filtre ─────────────────────────────────────────────────
    ws.freeze_panes = "B2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
