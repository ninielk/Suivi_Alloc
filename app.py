import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import date
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import plotly.graph_objects as go
import warnings

warnings.filterwarnings("ignore")

st.set_page_config(page_title="Suivi d'Allocation", layout="wide")

# ══════════════════════════════════════════════════════════════════
#  STRUCTURE FIXE — ROW_DEFS
#  (label, type, detail, B=critère AK, C=critère F, alloc_cible_fixe, C_knl)
#
#  detail :
#    normal            → calcul depuis Portefeuille (AK + F)
#    nantissement_souv → calcul depuis Nantissement (F + AK) + retraitement
#    nantissement_priv → calcul depuis Nantissement (F + OPCVM MONETAIRES)
#
#  alloc_cible_fixe :
#    float  → hardcodé (ex: 0.08 pour 8%)
#    None   → calculé dynamiquement (lignes 0-5)
# ══════════════════════════════════════════════════════════════════
ROW_DEFS = [
    # idx  label                                    type              detail                  B (AK)                                     C (F)                              alloc_fixe  C_knl
    # 0
    ("Obligations classiques", "blue", None, None, None, None, None),
    # 1
    ("Obligations souveraines", "white", "normal", "Obligations souveraines", "EMPRUNTS ETATS & OBLIG GARANTIES", None,
     None),
    # 2
    ("Obligations privées", "white", "normal", "Obligations privées", "OBLIGATIONS COTEES", None, None),
    # 3
    ("Obligations nanties", "nanties_parent", None, None, None, None, None),
    # 4
    ("Obligations souveraines", "white", "nantissement_souv", "Obligations souveraines",
     "EMPRUNTS ETATS & OBLIG GARANTIES", None, None),
    # 5
    ("Obligations privées", "white", "nantissement_priv", "Obligations privées", "OBLIGATIONS COTEES", None, None),
    # 6
    ("Autres produits de taux", "blue", None, None, None, 0.13, None),
    # 7
    ("Dettes privées", "white", "normal", "Dettes privées", None, 0.08, "DETTE PRIVEE"),
    # 8
    ("Alternatifs", "white", "normal", "Alternatifs", None, 0.05, "ALTERNATIF"),
    # 9
    ("Actions", "blue", None, None, None, 0.11, None),
    # 10
    ("Actions internationales", "white", "normal", "Actions internationales", None, 0.00, None),
    # 11
    ("Actions Zone Euro", "white", "normal", "Actions Zone Euro", None, 0.06, "ACTIONS"),
    # 12
    ("Autres actions (capital investissement)", "white", "normal", "Autres actions (capital investissement)", None,
     0.05, "Private Equity"),
    # 13
    ("Actifs réels", "blue", None, None, None, 0.19, None),
    # 14
    ("Immobilier placement", "white", "normal", "Immobilier placement", None, 0.13, "IMMOBILIER"),
    # 15
    ("Infrastructures", "white", "normal", "Infrastructures", None, 0.06, "INFRASTRUCTURE"),
    # 16
    ("Stratégique", "blue", None, None, None, 0.12, None),
    # 17
    ("Prêts stratégiques", "white", "normal", "Prêts stratégiques", None, 0.02, None),
    # 18
    ("Immobilier stratégique", "white", "normal", "Immobilier stratégique", None, 0.01, None),
    # 19
    ("Actions stratégiques", "white", "normal", "Actions stratégiques", None, 0.09, None),
    # 20
    ("Trésorerie", "blue", None, None, None, 0.01, None),
    # 21
    ("Trésorerie", "white", "normal", "Trésorerie", None, 0.01, None),
    # 22
    ("Total Général", "total", None, None, None, 1.00, None),
]

BLUE_CHILDREN = {
    0: [1, 2],
    3: [4, 5],
    6: [7, 8],
    9: [10, 11, 12],
    13: [14, 15],
    16: [17, 18, 19],
    20: [21],
}
BLUE_IDX = [0, 3, 6, 9, 13, 16, 20]
TOTAL_IDX = 22

# Marge de manoeuvre hardcodée (None = vide)
MARGE = {
    0: "-20% / +5%", 1: "-20% / +5%", 2: "-20% / +5%",
    6: "-13% / +3%", 7: "-8% / +3%", 8: "-5% / +3%",
    9: "-11% / +3%", 10: "-0% / +3%", 11: "-6% / +3%", 12: "-5% / +3%",
    13: "-19% / +3%", 14: "-13% / +3%", 15: "-6% / +3%",
}

REQUIRED_SHEETS = ["Portefeuille", "Retraitements", "Nantissement"]


# ══════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════
def normalize(s):
    import unicodedata
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return " ".join(s.split()).upper()


def somme_si(df, crit_col, crit_val, sum_col):
    if crit_val is None or df.empty:
        return 0.0
    if isinstance(crit_val, float) and np.isnan(crit_val):
        return 0.0
    target = normalize(crit_val)
    mask = df[crit_col].astype(str).map(normalize) == target
    return float(df.loc[mask, sum_col].sum())


def fmt_m(val, decimals=0):
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return ""
    return f"{val:,.{decimals}f}"


def fmt_pct(val, decimals=1):
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return ""
    return f"{val * 100:.{decimals}f}%"


# ══════════════════════════════════════════════════════════════════
#  CALCUL
# ══════════════════════════════════════════════════════════════════
@st.cache_data(show_spinner=False)
def compute(file_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

    # ── Noms de colonnes attendus (robuste aux insertions de colonnes)
    # Portefeuille/Nantissement : on cherche par fragment de nom
    PORT_COLS = {
        "F": ["catégorie d'instrument", "categorie d instrument", "catégorie instrument"],
        "W": ["valeur actuelle comptable", "valeur comptable", "vnc"],
        "Y": ["valeur de marché", "valeur marche", "val. marché hors"],
        "AK": ["classification", "classe d'actif", "classe actif"],
    }
    NANT_COLS = {
        "F": ["catégorie d'instrument", "categorie d instrument", "catégorie instrument"],
        "W": ["valeur actuelle comptable", "valeur comptable", "vnc"],
        "Y": ["valeur de marché", "valeur marche", "val. marché hors"],
    }

    def find_col(headers_lower, candidates):
        """Trouve l'index d'une colonne par fragment de nom (case-insensitive)."""
        for i, h in enumerate(headers_lower):
            if h is None: continue
            for c in candidates:
                if c in h:
                    return i
        return None

    def read_by_header(ws, col_map, warnings_list, sheet_name):
        """Lit un onglet en cherchant les colonnes par leur nom."""
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers_lower = [str(h).lower().strip() if h else "" for h in header_row]

        idx = {}
        for key, candidates in col_map.items():
            found = find_col(headers_lower, candidates)
            if found is None and key != "AK":  # AK optionnel pour Nantissement
                warnings_list.append(f"⚠️ Colonne '{key}' introuvable dans **{sheet_name}** — vérifiez les headers.")
            idx[key] = found

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row: continue
            r = {}
            for key, i in idx.items():
                r[key] = row[i] if (i is not None and i < len(row)) else None
            rows.append(r)

        df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=list(col_map.keys()))
        for col in ["W", "Y"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        if "AK" not in df.columns:
            df["AK"] = None
        return df

    # Lecture Portefeuille & Nantissement par header
    _warnings = []
    df_p = read_by_header(wb["Portefeuille"], PORT_COLS, _warnings, "Portefeuille")
    df_nant = read_by_header(wb["Nantissement"], NANT_COLS, _warnings,
                             "Nantissement") if "Nantissement" in wb.sheetnames else pd.DataFrame(
        columns=["F", "W", "Y", "AK"])
    if "AK" not in df_nant.columns: df_nant["AK"] = None

    # (Portefeuille & Nantissement déjà lus par read_by_header ci-dessus)

    # ── Retraitements : C=2 classe, D=3 montant
    ws_r, rows_r = wb["Retraitements"], []
    for row in ws_r.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 4:
            continue
        classe, montant = row[2], row[3]
        if classe is None or str(classe).strip().lower() in ("classe d'actifs", "montant", ""):
            continue
        try:
            rows_r.append({"classe": str(classe).strip(), "D": float(montant)})
        except (TypeError, ValueError):
            continue
    df_r = pd.DataFrame(rows_r) if rows_r else pd.DataFrame(columns=["classe", "D"])

    # ── KNL : B=1, K=10, M=12, N=13 + date E1
    has_knl = "KNL" in wb.sheetnames
    df_knl = pd.DataFrame(columns=["B", "K", "M", "N"])
    knl_year = None
    if has_knl:
        ws_knl = wb["KNL"]
        # Lecture date KNL!E1 (col index 4)
        knl_e1 = ws_knl.cell(row=1, column=5).value
        if hasattr(knl_e1, "year"):
            knl_year = knl_e1.year
        rows_knl = []
        for row in ws_knl.iter_rows(min_row=3, values_only=True):
            if not row or len(row) <= 13 or row[1] is None:
                continue
            rows_knl.append({"B": row[1], "K": row[10], "M": row[12], "N": row[13]})
        if rows_knl:
            df_knl = pd.DataFrame(rows_knl)
            for col in ["K", "M", "N"]:
                df_knl[col] = pd.to_numeric(df_knl[col], errors="coerce").fillna(0)

    # ── Calcul L8 (engagements nanties privées)
    l8 = 0.0
    if knl_year is not None:
        l8 = -35.0 * (knl_year - date.today().year)

    # ── Onglet Params : durées ALM et chocs SCR modifiables par Jérôme
    # Valeurs par défaut si onglet absent
    PARAMS_DEFAULT = {
        "Actions internationales": ("dur_alm", 20.0),
        "Actions Zone Euro": ("dur_alm", 20.0),
        "Autres actions (capital investissement)": ("dur_alm", 10.0),
        "Immobilier placement": ("dur_alm", 20.0),
        "Infrastructures": ("dur_alm", 20.0),
        "Prêts stratégiques": ("dur_alm", 10.0),
        "Immobilier stratégique": ("dur_alm", 20.0),
        "Actions stratégiques": ("dur_alm", 10.0),
        "Trésorerie": ("dur_alm", 1.0),
        "Choc SCR Immobilier": ("scr", 0.22),
        "Choc SCR Infrastructures": ("scr", 0.325),
    }
    params = {k: v[1] for k, v in PARAMS_DEFAULT.items()}

    if "Params" in wb.sheetnames:
        ws_params = wb["Params"]
        for row in ws_params.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None or row[1] is None:
                continue
            key = str(row[0]).strip()
            try:
                val = float(row[1])
                if key in params:
                    params[key] = val
            except (TypeError, ValueError):
                continue
    else:
        _warnings.append("ℹ️ Onglet **Params** absent — durées ALM et chocs SCR par défaut utilisés.")

    # Mapping label → idx Alloc pour U_MANUAL
    LABEL_TO_IDX = {
        "Actions internationales": 10,
        "Actions Zone Euro": 11,
        "Autres actions (capital investissement)": 12,
        "Immobilier placement": 14,
        "Infrastructures": 15,
        "Prêts stratégiques": 17,
        "Immobilier stratégique": 18,
        "Actions stratégiques": 19,
        "Trésorerie": 21,
    }
    U_MANUAL_PARAMS = {LABEL_TO_IDX[k]: params[k] for k in LABEL_TO_IDX}
    SCR_IMMO_PARAMS = params.get("Choc SCR Immobilier", 0.22)
    SCR_INFRA_PARAMS = params.get("Choc SCR Infrastructures", 0.325)

    # ── TPT : lecture pour calculs SCR et duration
    #    I=8 NAV | Z=25 valeur marché | CT=97 fonds | EQ=146 duration
    #    CU=98 SCR taux | CW=100 SCR eq1 | CX=101 SCR eq2
    #    CY=102 SCR immo | CZ=103 SCR spread | DE=108 SCR forex | EX=153 CIC
    has_tpt = "TPT" in wb.sheetnames
    tpt_scr = {}  # résultats SCR par catégorie CIC
    tpt_dur = {}  # duration pondérée
    tpt_nav = 0.0
    tpt_dur_global = 0.0
    tpt_dur_souv = 0.0
    tpt_dur_entrep = 0.0
    tpt_dur_autres = 0.0

    if has_tpt:
        ws_tpt = wb["TPT"]
        # Lecture TPT par nom de colonne — robuste aux insertions
        TPT_COL_NAMES = {
            "NAV": "8b_total_number_of_shares",
            "Z": "26_net_asset_value",
            "CT": "95_identification_of_the_original_portfolio",
            "EQ": "144_modified_duration",
            "CU": "97_scr_mrkt_ir_up",
            "CW": "99_scr_mrkt_eq_type1",
            "CX": "100_scr_mrkt_eq_type2",
            "CY": "101_scr_mrkt_prop",
            "CZ": "102_scr_mrkt_spread",
            "DE": "105b_scr_mrkt_fx",
            "EX": "cic 1",  # colonne EX = "CIC 1"
        }
        tpt_header_row = next(ws_tpt.iter_rows(min_row=1, max_row=1, values_only=True))
        tpt_headers_lower = [str(h).lower().strip() if h else "" for h in tpt_header_row]

        tpt_idx = {}
        for key, fragment in TPT_COL_NAMES.items():
            for i, h in enumerate(tpt_headers_lower):
                if fragment in h:
                    tpt_idx[key] = i
                    break
            if key not in tpt_idx:
                _warnings.append(f"⚠️ Colonne TPT '{key}' ({fragment}) introuvable — SCR/Duration peut être incorrect.")
                tpt_idx[key] = None

        rows_tpt = []

        def _n(v):
            return float(v) if v is not None else 0.0

        for row in ws_tpt.iter_rows(min_row=2, values_only=True):
            if not row: continue

            def g(key):
                i = tpt_idx.get(key)
                return row[i] if (i is not None and i < len(row)) else None

            rows_tpt.append({
                "NAV": _n(g("NAV")),
                "Z": _n(g("Z")),
                "CT": g("CT"),
                "EQ": _n(g("EQ")),
                "CU": _n(g("CU")),
                "CW": _n(g("CW")),
                "CX": _n(g("CX")),
                "CY": _n(g("CY")),
                "CZ": _n(g("CZ")),
                "DE": _n(g("DE")),
                "EX": g("EX"),
            })
        if rows_tpt:
            df_tpt = pd.DataFrame(rows_tpt)
            df_tpt["EX_str"] = df_tpt["EX"].astype(str).str.strip()
            df_tpt["CT_vide"] = df_tpt["CT"].isna() | (df_tpt["CT"].astype(str).str.strip() == "")

            # NAV = valeur col I (même sur chaque ligne)
            tpt_nav = df_tpt["NAV"].iloc[0] if len(df_tpt) > 0 else 1.0
            if tpt_nav == 0:
                tpt_nav = 1.0

            # ── Durations pondérées (transparisé = CT vide)
            direct = df_tpt[df_tpt["CT_vide"]]
            z_sum_all = direct["Z"].sum()

            # Souveraines (CIC commençant par 1)
            souv = direct[direct["EX_str"].str.startswith("1")]
            z_s = souv["Z"].sum()
            tpt_dur_souv = (souv["EQ"] * souv["Z"]).sum() / z_s if z_s else 0.0

            # Entreprises (CIC commençant par 2)
            entr = direct[direct["EX_str"].str.startswith("2")]
            z_e = entr["Z"].sum()
            tpt_dur_entrep = (entr["EQ"] * entr["Z"]).sum() / z_e if z_e else 0.0

            # Autres produits de taux (CT non vide = transparisé)
            autres = df_tpt[~df_tpt["CT_vide"]]
            z_a = autres["Z"].sum()
            tpt_dur_autres = (autres["EQ"] * autres["Z"]).sum() / z_a if z_a else 0.0

            # Duration globale S25 = SOMMEPROD(EQ:EQ; Z:Z)/SOMME(Z:Z) sur TOUTES les lignes
            z_all = df_tpt["Z"].sum()
            tpt_dur_global = (df_tpt["EQ"] * df_tpt["Z"]).sum() / z_all if z_all else 0.0

            corr = np.array([
                [1, 0, 0, 0, 0.25],
                [0, 1, 0.75, 0.5, 0.25],
                [0, 0.75, 1, 0.75, 0.25],
                [0, 0.5, 0.75, 1, 0.25],
                [0.25, 0.25, 0.25, 0.25, 1],
            ])

            # ── SCR par CIC
            cic_list = ["1", "2", "3", "4", "5", "7", "8", "9", "A", "B", "D", "E"]
            for cic in cic_list:
                sub = df_tpt[df_tpt["EX_str"] == cic]
                z = sub["Z"].sum()
                scr_taux = sub["CU"].sum() * tpt_nav
                scr_spread = sub["CZ"].sum() * tpt_nav
                scr_eq1 = sub["CW"].sum() * tpt_nav
                scr_eq2 = sub["CX"].sum() * tpt_nav
                scr_immo = sub["CY"].sum() * tpt_nav
                scr_forex = sub["DE"].sum() * tpt_nav
                scr_actions = np.sqrt(scr_eq1 ** 2 + scr_eq2 ** 2 + 0.75 * scr_eq1 * scr_eq2) if (
                            scr_eq1 or scr_eq2) else 0.0
                v = np.array([scr_taux, scr_spread, scr_actions, scr_immo, scr_forex])
                scr_total = float(np.sqrt(v @ corr @ v)) if v.any() else 0.0
                dur_w = (sub["EQ"] * sub["Z"]).sum() / z if z else 0.0
                tpt_scr[cic] = {
                    "FA": z / 1e6,
                    "FB": dur_w,
                    "FC": scr_total / 1e6,
                    "FD": scr_taux / 1e6,
                    "FE": scr_spread / 1e6,
                    "FF": scr_actions / 1e6,
                    "FG": scr_immo / 1e6,
                    "FH": scr_forex / 1e6,
                    "FK": scr_total / z if z else 0.0,
                }

            # ── SCR TOTAL : calculé sur le vecteur global (diversification correcte)
            fa21 = df_tpt["Z"].sum() / 1e6
            scr_taux_tot = df_tpt["CU"].sum() * tpt_nav
            scr_spread_tot = df_tpt["CZ"].sum() * tpt_nav
            scr_eq1_tot = df_tpt["CW"].sum() * tpt_nav
            scr_eq2_tot = df_tpt["CX"].sum() * tpt_nav
            scr_immo_tot = df_tpt["CY"].sum() * tpt_nav
            scr_forex_tot = df_tpt["DE"].sum() * tpt_nav
            scr_act_tot = np.sqrt(scr_eq1_tot ** 2 + scr_eq2_tot ** 2 + 0.75 * scr_eq1_tot * scr_eq2_tot) if (
                        scr_eq1_tot or scr_eq2_tot) else 0.0
            v_tot = np.array([scr_taux_tot, scr_spread_tot, scr_act_tot, scr_immo_tot, scr_forex_tot])
            fc21 = float(np.sqrt(v_tot @ corr @ v_tot)) / 1e6 if v_tot.any() else 0.0
            tpt_scr["TOTAL"] = {
                "FA": fa21,
                "FC": fc21,
                "FD": scr_taux_tot / 1e6,
                "FK": (fc21 / fa21) if fa21 else 0.0,
            }

    # ══════════════════════════════════════════════
    #  ETAPE 1 : calcul nanties EN PREMIER (nécessaires pour classiques)
    # ══════════════════════════════════════════════
    res = {}

    # idx 4 — Oblig souveraines nanties
    # G7 = SOMME.SI(Nant!F, C7, Y) + SOMME.SI(Nant!AK, B7, Y)
    C4 = ROW_DEFS[4][4]  # = "EMPRUNTS ETATS & OBLIG GARANTIES"
    B4 = ROW_DEFS[4][3]  # = "Obligations souveraines"
    g4 = (somme_si(df_nant, "F", C4, "Y") + somme_si(df_nant, "AK", B4, "Y")) / 1e6
    j4 = (somme_si(df_nant, "F", C4, "W") + somme_si(df_nant, "AK", B4, "W")) / 1e6
    res[4] = {"label": ROW_DEFS[4][0], "type": "white", "detail": "nantissement_souv",
              "D": 0.0, "G": g4, "J": j4, "M": 0.0, "N": 0.0, "O": 0.0}

    # idx 5 — Oblig privées nanties
    # G8 = SOMME.SI(Nant!F, C8, Y) + SOMME.SI(Nant!F, "OPCVM MONETAIRES", Y)
    C5 = ROW_DEFS[5][4]  # = "OBLIGATIONS COTEES"
    B5 = ROW_DEFS[5][3]  # = "Obligations privées"
    g5 = (somme_si(df_nant, "F", C5, "Y") + somme_si(df_nant, "F", "OPCVM MONETAIRES", "Y")) / 1e6
    j5 = (somme_si(df_nant, "F", C5, "W") + somme_si(df_nant, "AK", B5, "W")) / 1e6
    res[5] = {"label": ROW_DEFS[5][0], "type": "white", "detail": "nantissement_priv",
              "D": 0.0, "G": g5, "J": j5, "M": l8, "N": 0.0, "O": 0.0}

    # ══════════════════════════════════════════════
    #  ETAPE 2 : toutes les autres lignes blanches
    # ══════════════════════════════════════════════
    for idx, (label, rtype, detail, B, C, _, C_knl) in enumerate(ROW_DEFS):
        if rtype != "white" or idx in (4, 5):
            continue

        d = 0.0;
        g = 0.0;
        j = None;
        m_val = 0.0;
        n_val = 0.0;
        o_val = 0.0

        if detail == "normal":
            d = somme_si(df_r, "classe", B, "D") / 1e6
            g_raw = (somme_si(df_p, "AK", B, "Y") + (somme_si(df_p, "F", C, "Y") if C else 0.0)) / 1e6
            j_raw = (somme_si(df_p, "AK", B, "W") + (somme_si(df_p, "F", C, "W") if C else 0.0)) / 1e6

            # Obligations souveraines classiques (idx=1) : soustraire nanties souv
            if idx == 1:
                g = g_raw + d - res[4]["G"]
                j = j_raw - res[4]["J"]
            # Obligations privées classiques (idx=2) : soustraire nanties priv
            elif idx == 2:
                g = g_raw + d - res[5]["G"]
                j = j_raw - res[5]["J"]
            else:
                g = g_raw + d
                j = j_raw

            m_val = somme_si(df_knl, "B", C_knl, "K") / 1e6 if C_knl else 0.0
            n_val = somme_si(df_knl, "B", C_knl, "M") / 1e6 if C_knl else 0.0
            o_val = somme_si(df_knl, "B", C_knl, "N") / 1e6 if C_knl else 0.0

        res[idx] = {"label": label, "type": rtype, "detail": detail,
                    "D": d, "G": g, "J": j, "M": m_val, "N": n_val, "O": o_val}

    # ══════════════════════════════════════════════
    #  ETAPE 3 : lignes bleues = somme enfants
    # ══════════════════════════════════════════════
    for br, children in BLUE_CHILDREN.items():
        rtype_br = ROW_DEFS[br][1]
        g_sum = sum(res[c]["G"] for c in children if c in res)
        j_vals = [res[c]["J"] for c in children if c in res and res[c]["J"] is not None]
        j_sum = sum(j_vals) if j_vals else None
        m_sum = sum(res[c]["M"] for c in children if c in res)
        n_sum = sum(res[c]["N"] for c in children if c in res)
        o_sum = sum(res[c]["O"] for c in children if c in res)
        res[br] = {"label": ROW_DEFS[br][0], "type": rtype_br, "detail": None,
                   "D": 0.0, "G": g_sum, "J": j_sum,
                   "M": m_sum, "N": n_sum, "O": o_sum}

    # ══════════════════════════════════════════════
    #  ETAPE 4 : cas spéciaux N + Total
    # ══════════════════════════════════════════════

    # N[Trésorerie] EN PREMIER : N24 = G24 - G_total * alloc_cible_tres
    g_total_tmp = sum(res[i]["G"] for i in BLUE_IDX)
    res[21]["N"] = res[21]["G"] - g_total_tmp * 0.01
    res[20]["N"] = res[21]["N"]

    # N[Oblig privées classiques] : N5 = -N9-N12-N16-N19-N23 + M25 - M6(nanties)
    m_total = sum(res[i]["M"] for i in BLUE_IDX)  # = M25
    n_autres = sum(res[i]["N"] for i in [6, 9, 13, 16, 20])
    m_nanties = res[3]["M"]  # = M6
    res[2]["N"] = -n_autres + m_total - m_nanties
    res[0]["N"] = res[2]["N"]  # N3 = N5

    # Total général
    g_tot = sum(res[i]["G"] for i in BLUE_IDX)
    j_vals_tot = [res[i]["J"] for i in BLUE_IDX if res[i].get("J") is not None]
    j_tot = sum(j_vals_tot) if j_vals_tot else 0.0
    m_tot = sum(res[i]["M"] for i in BLUE_IDX)
    n_tot = sum(res[i]["N"] for i in BLUE_IDX)
    o_tot = sum(res[i]["O"] for i in BLUE_IDX)
    res[TOTAL_IDX] = {"label": "Total Général", "type": "total", "detail": None,
                      "D": 0.0, "G": g_tot, "J": j_tot,
                      "M": m_tot, "N": n_tot, "O": o_tot}

    # ══════════════════════════════════════════════
    #  ETAPE 5 : alloc cible dynamique
    # ══════════════════════════════════════════════
    # E6=G6/G25, E7=G7/G25, E8=G8/G25
    # E3=44%-E6, E4=20%-E7, E5=24%-E8
    e6 = res[3]["G"] / g_tot if g_tot else 0.0
    e7 = res[4]["G"] / g_tot if g_tot else 0.0
    e8 = res[5]["G"] / g_tot if g_tot else 0.0
    res[3]["alloc_cible_f"] = e6
    res[4]["alloc_cible_f"] = e7
    res[5]["alloc_cible_f"] = e8
    res[0]["alloc_cible_f"] = 0.44 - e6
    res[1]["alloc_cible_f"] = 0.20 - e7
    res[2]["alloc_cible_f"] = 0.24 - e8
    for idx, (_, _, _, _, _, alloc_fixe, _) in enumerate(ROW_DEFS):
        if idx in res and "alloc_cible_f" not in res[idx]:
            res[idx]["alloc_cible_f"] = alloc_fixe

    # ══════════════════════════════════════════════
    #  ETAPE 6 : pourcentages et dérivés
    # ══════════════════════════════════════════════

    # H = G% , K = J%
    for r in res.values():
        r["H"] = r["G"] / g_tot if g_tot else 0.0
        r["K"] = (r["J"] / j_tot) if (j_tot and r.get("J") is not None) else None

    # +/- values = G - J, arrondi pour eviter erreurs floating point
    for r in res.values():
        j_val = r.get("J")
        if j_val is not None:
            raw = r["G"] - j_val
            r["PLUSMOINS"] = round(raw, 6) if abs(raw) > 1e-4 else 0.0
        else:
            r["PLUSMOINS"] = None

    # Allocation projetée P = G + M - N + O
    for idx, r in res.items():
        r["P"] = r["G"] + r["M"] - r["N"] + r["O"]
    # Nanties parent = somme enfants
    res[3]["P"] = sum(res[c]["P"] for c in BLUE_CHILDREN[3] if c in res)

    # Q = P%
    # P3 = O3/O25 - P6, P4 = O4/O25 - P7, P5 = O5/O25 - P8
    p_tot = res[TOTAL_IDX]["P"]
    for idx, r in res.items():
        if p_tot:
            q_raw = r["P"] / p_tot
            if idx == 0:
                r["Q"] = q_raw - res[3]["P"] / p_tot
            elif idx == 1:
                r["Q"] = q_raw - res[4]["P"] / p_tot
            elif idx == 2:
                r["Q"] = q_raw - res[5]["P"] / p_tot
            else:
                r["Q"] = q_raw
        else:
            r["Q"] = None

    # ══════════════════════════════════════════════
    #  ETAPE 7 : colonnes S, U, T, V, W, X, Y, ECART, R
    # ══════════════════════════════════════════════

    fk = {cic: tpt_scr.get(cic, {}).get("FK", 0.0) for cic in ["1", "2", "3", "4", "7", "8", "9", "A", "B", "D", "E"]}
    fk_total = tpt_scr.get("TOTAL", {}).get("FK", 0.0)
    fk_cic4 = fk.get("4", 0.0)

    # ── Colonne S : duration par catégorie depuis TPT
    # S4 = TPT!FA2 = tpt_dur_souv, S5 = TPT!FA3 = tpt_dur_entrep
    # S7=S4, S8=S5, S10=S11=TPT!FA4=tpt_dur_autres
    # S20 (Excel row 20 = idx 17 Prêts strat) = S8 = tpt_dur_entrep
    # S25 = tpt_dur_global
    S_MAP = {
        1: tpt_dur_souv, 2: tpt_dur_entrep,
        4: tpt_dur_souv, 5: tpt_dur_entrep,
        7: tpt_dur_autres, 8: tpt_dur_autres,
        17: tpt_dur_entrep,  # S20=S8
    }
    for idx, r in res.items():
        r["S"] = tpt_dur_global if idx == TOTAL_IDX else S_MAP.get(idx, 0.0)

    # ── Colonne U : duration ALM (S pour rows 1-8, manuelles pour reste)
    # U25 = SOMMEPROD(U3:U24 ; H3:H24)
    U_MANUAL = U_MANUAL_PARAMS
    # U4-U11 = S4-S11 (idx 1-8 seulement, pas idx 17 qui est manuel)
    U_FROM_S = {1, 2, 4, 5, 7, 8}
    for idx, r in res.items():
        if idx == TOTAL_IDX: continue
        if idx in U_FROM_S:
            r["U"] = S_MAP.get(idx, 0.0)
        elif idx in U_MANUAL:
            r["U"] = U_MANUAL[idx]
        else:
            r["U"] = 0.0
    # U25 = SOMMEPROD(U ; H)
    u25 = sum((res[i].get("U") or 0.0) * (res[i].get("H") or 0.0)
              for i in range(len(ROW_DEFS)) if i in res and i != TOTAL_IDX)
    res[TOTAL_IDX]["U"] = u25

    # ── Colonne X : SCR/actifs par catégorie (FK depuis TPT)
    X_MAP = {
        1: fk.get("1", 0.0), 2: fk.get("2", 0.0),
        4: fk.get("1", 0.0), 5: fk.get("2", 0.0),
        7: fk.get("8", 0.0), 8: fk.get("2", 0.0),
        10: fk_cic4, 11: fk_cic4, 12: fk_cic4 + 0.10,
        14: 0.22, 15: 0.325,
    }
    for idx, r in res.items():
        r["X"] = X_MAP.get(idx, 0.0) if idx != TOTAL_IDX else 0.0

    # ── Colonne W : SCR/actifs global — seulement sur le total
    for idx, r in res.items():
        r["W"] = fk_total if idx == TOTAL_IDX else None

    # ── Colonnes T et V : contributions duration
    # T_i = S_i × (P_i% - H_i%) | T25 = S25 + SOMME(T4:T24)
    # V_i = U_i × (P_i% - H_i%) | V25 = U25 + SOMME(V4:V24)
    for idx, r in res.items():
        if idx == TOTAL_IDX: continue
        p_pct = r.get("Q") or 0.0
        h_pct = r.get("H") or 0.0
        delta = p_pct - h_pct
        r["T"] = (r.get("S") or 0.0) * delta
        r["V"] = (r.get("U") or 0.0) * delta
    sum_t = sum(r.get("T") or 0.0 for i, r in res.items() if i != TOTAL_IDX)
    sum_v = sum(r.get("V") or 0.0 for i, r in res.items() if i != TOTAL_IDX)
    res[TOTAL_IDX]["T"] = tpt_dur_global + sum_t  # T25 = S25 + SOMME(T)
    res[TOTAL_IDX]["V"] = u25 + sum_v  # V25 = U25 + SOMME(V)

    # ── Colonne Y : contribution SCR taux
    # Y_i = X_i × (P_i% - H_i%) | Y25 = W25 + SOMME(Y4:Y24)
    for idx, r in res.items():
        if idx == TOTAL_IDX: continue
        p_pct = r.get("Q") or 0.0
        h_pct = r.get("H") or 0.0
        r["Y_col"] = (r.get("X") or 0.0) * (p_pct - h_pct)
    sum_y = sum(r.get("Y_col") or 0.0 for i, r in res.items() if i != TOTAL_IDX)
    res[TOTAL_IDX]["Y_col"] = fk_total + sum_y  # Y25 = W25 + SOMME(Y)

    # ── Colonne ECART (Excel col Q) et R
    # Q3=E3-P3-P6, Q9=E9-P9, Q12=E12-P12, Q16=E16-P16, Q19=E19-P19, Q23=E23-P23
    # (PAS de ECART pour nanties idx 3)
    for idx, r in res.items():
        ac = r.get("alloc_cible_f") or 0.0
        q = r.get("Q") or 0.0
        if idx == 0:  # Q3 = E3 - P3 - P6
            r["ECART"] = ac - q - (res[3].get("Q") or 0.0)
        elif idx in (6, 9, 13, 16, 20):  # Q9, Q12, Q16, Q19, Q23
            r["ECART"] = ac - q
        else:
            r["ECART"] = None
    p_tot_r = res[TOTAL_IDX].get("P") or 0.0
    for r in res.values():
        ecart = r.get("ECART")
        r["R"] = (ecart * p_tot_r) if ecart is not None else None

    return res, _warnings


# ══════════════════════════════════════════════════════════════════
#  STYLE
# ══════════════════════════════════════════════════════════════════
# Couleurs DA LMG
LMG_COLORS = ["#B10967", "#412761", "#007078", "#F8AF00", "#99DBF2", "#D3E8CA", "#6E6E6E", "#243D7C", "#E85D75",
              "#5B9BD5", "#70AD47", "#FFC000"]

STYLE = {
    "blue": {"bg": "#B10967", "fg": "#FFFFFF", "fw": "700"},
    "nanties_parent": {"bg": "#E85D75", "fg": "#FFFFFF", "fw": "700"},
    "white": {"bg": "#FFFFFF", "fg": "#412761", "fw": "400"},
    "total": {"bg": "#412761", "fg": "#FFFFFF", "fw": "700"},
}

MARGE = {
    0: "-20% / +5%", 1: "-20% / +5%", 2: "-20% / +5%",
    6: "-13% / +3%", 7: "-8% / +3%", 8: "-5% / +3%",
    9: "-11% / +3%", 10: "-0% / +3%", 11: "-6% / +3%", 12: "-5% / +3%",
    13: "-19% / +3%", 14: "-13% / +3%", 15: "-6% / +3%",
}


def fmt_m(val, decimals=0):
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return ""
    return f"{val:,.{decimals}f}"


def fmt_pct(val, decimals=1):
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return ""
    return f"{val * 100:.{decimals}f}%"


# ══════════════════════════════════════════════════════════════════
#  RENDU HTML
# ══════════════════════════════════════════════════════════════════
def render_table(res: dict, show_retraitement: bool = False, show_qy: bool = False) -> str:
    headers = [
        ("Catégorie", "l", True),
        ("Cible", "c", True),
        ("Marge", "c", True),
        ("Retrait. M€", "r", show_retraitement),
        ("Alloc. M€", "r", True),
        ("Alloc. %", "r", True),
        ("VNC M€", "r", True),
        ("VNC %", "r", True),
        ("+/- val. M€", "r", True),
        ("Engag. M€", "r", True),
        ("Fin. appels M€", "r", True),
        ("Gains cap. M€", "r", True),
        ("Proj. M€", "r", True),
        ("Proj. %", "r", True),
        ("Écart cible %", "r", show_qy),
        ("Écart cible M€", "r", show_qy),
        ("Duration", "r", show_qy),
        ("Contrib. dur.", "r", show_qy),
        ("Duration ALM", "r", show_qy),
        ("Contrib. dur. ALM", "r", show_qy),
        ("SCR/Actifs global", "r", show_qy),
        ("SCR/Actifs cat.", "r", show_qy),
        ("Contrib. SCR taux", "r", show_qy),
    ]
    th = "".join(f'<th class="{a}">{h}</th>' for h, a, show in headers if show)
    html = f"""
    <style>
      .at{{border-collapse:collapse;width:100%;font-family:Calibri,sans-serif;font-size:11px}}
      .at th{{padding:5px 6px;text-align:center;border:1px solid #888;
             background:#412761;color:#fff;font-weight:700;white-space:nowrap}}
      .at td{{padding:3px 6px;border:1px solid #ccc;white-space:nowrap}}
      .r{{text-align:right}}.l{{text-align:left}}.c{{text-align:center}}
    </style>
    <table class="at"><thead><tr>{th}</tr></thead><tbody>
    """
    for idx, (label, rtype, detail, _, _, _, _) in enumerate(ROW_DEFS):
        r = res.get(idx)
        if not r:
            continue
        s = STYLE[rtype]
        cs = f'background:{s["bg"]};color:{s["fg"]};font-weight:{s["fw"]};'

        ac_f = r.get("alloc_cible_f")
        dur = r.get("S") or 0.0
        scr = r.get("X") or 0.0

        cells = [
            (label, "l", True),
            (fmt_pct(ac_f) if ac_f is not None else "", "c", True),
            (MARGE.get(idx, ""), "c", True),
            (fmt_m(r["D"]) if (rtype == "white" and detail in ("normal", "nantissement_souv")) else "", "r",
             show_retraitement),
            (fmt_m(r["G"]), "r", True),
            (fmt_pct(r["H"]), "r", True),
            (fmt_m(r["J"]) if r.get("J") is not None else "", "r", True),
            (fmt_pct(r["K"]) if r.get("K") is not None else "", "r", True),
            (fmt_m(r["PLUSMOINS"]) if r.get("PLUSMOINS") is not None else "", "r", True),
            (fmt_m(r["M"]) if r.get("M") else "", "r", True),
            (fmt_m(r["N"]) if r.get("N") else "", "r", True),
            (fmt_m(r["O"]) if r.get("O") else "", "r", True),
            (fmt_m(r["P"]) if r.get("P") is not None else "", "r", True),
            (fmt_pct(r["Q"]) if r.get("Q") is not None else "", "r", True),
            (fmt_pct(r.get("ECART")) if r.get("ECART") is not None else "", "r", show_qy),
            (fmt_m(r.get("R")) if r.get("R") is not None else "", "r", show_qy),
            (f"{(r.get('S') or 0):.2f}" if r.get("S") else "", "r", show_qy),
            (f"{(r.get('T') or 0):.3f}" if r.get("T") else "", "r", show_qy),
            (f"{(r.get('U') or 0):.2f}" if r.get("U") else "", "r", show_qy),
            (f"{(r.get('V') or 0):.3f}" if r.get("V") else "", "r", show_qy),
            (f"{(r.get('W') or 0) * 100:.2f}%" if (r.get("W") and idx == TOTAL_IDX) else "", "r", show_qy),
            (f"{(r.get('X') or 0) * 100:.2f}%" if r.get("X") else "", "r", show_qy),
            (f"{(r.get('Y_col') or 0) * 100:.2f}%" if r.get("Y_col") else "", "r", show_qy),
        ]
        tds = "".join(f'<td class="{a}" style="{cs}">{v}</td>' for v, a, show in cells if show)
        html += f'<tr>{tds}</tr>\n'
    html += "</tbody></table>"
    return html


# ══════════════════════════════════════════════════════════════════
#  EXPORT EXCEL
# ══════════════════════════════════════════════════════════════════
def export_excel(res: dict) -> BytesIO:
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alloc"
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    FILLS = {
        "blue": PatternFill(fgColor="B10967", fill_type="solid"),
        "nanties_parent": PatternFill(fgColor="E85D75", fill_type="solid"),
        "white": PatternFill(fgColor="FFFFFF", fill_type="solid"),
        "total": PatternFill(fgColor="412761", fill_type="solid"),
        "header": PatternFill(fgColor="412761", fill_type="solid"),
    }
    FONTS = {
        "blue": Font(bold=True, color="FFFFFF", name="Calibri", size=11),
        "nanties_parent": Font(bold=True, color="FFFFFF", name="Calibri", size=11),
        "white": Font(color="412761", name="Calibri", size=11),
        "total": Font(bold=True, color="FFFFFF", name="Calibri", size=11),
        "header": Font(bold=True, color="FFFFFF", name="Calibri", size=11),
    }
    headers = ["Catégorie", "Cible", "Marge", "Retraitement M€",
               "Alloc. M€", "Alloc. %", "VNC M€", "VNC %",
               "+/- val. M€", "Engag. M€", "Fin. appels M€", "Gains cap. M€",
               "Proj. M€", "Proj. %",
               "Écart cible %", "Écart cible M€", "Duration", "Contrib. dur.",
               "Duration ALM", "Contrib. dur. ALM", "SCR/Actifs global", "SCR/Actifs cat.", "Contrib. SCR taux"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = FILLS["header"];
        c.font = FONTS["header"];
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")
    for erow, (idx, (label, rtype, detail, _, _, _, _)) in enumerate(enumerate(ROW_DEFS), 2):
        r = res.get(idx)
        if not r:
            continue
        ac_f = r.get("alloc_cible_f")
        d_val = r["D"] if (rtype == "white" and detail in ("normal", "nantissement_souv")) else None
        vals = [label, ac_f, MARGE.get(idx, ""), d_val,
                r["G"], r["H"],
                r["J"] if r.get("J") is not None else None,
                r["K"] if r.get("K") is not None else None,
                r["PLUSMOINS"] if r.get("PLUSMOINS") is not None else None,
                r["M"] if r.get("M") else None,
                r["N"] if r.get("N") else None,
                r["O"] if r.get("O") else None,
                r["P"] if r.get("P") is not None else None,
                r["Q"] if r.get("Q") is not None else None,
                r.get("ECART") or None,
                r.get("R") or None,
                r.get("S") or None,
                r.get("T") or None,
                r.get("U") or None,
                r.get("V") or None,
                r.get("W") if idx == TOTAL_IDX else None,
                r.get("X") or None,
                r.get("Y_col") or None,
                ]
        fmts = [None, "0%", "@", "#,##0", "#,##0", "0.0%", "#,##0", "0.0%",
                "#,##0", "#,##0", "#,##0", "#,##0", "#,##0", "0.0%",
                "0.0%", "#,##0", "0.00", "0.000", "0.00", "0.000", "0.00%", "0.00%", "0.00%"]
        for col, (val, fmt) in enumerate(zip(vals, fmts), 1):
            c = ws.cell(row=erow, column=col, value=val)
            c.fill = FILLS[rtype];
            c.font = FONTS[rtype];
            c.border = border
            c.alignment = Alignment(horizontal="left" if col == 1 else "right", vertical="center")
            if fmt and val is not None and val != "":
                c.number_format = fmt
    ws.column_dimensions["A"].width = 38
    for col_letter in "BCDEFGHIJKLMNOPQRSTUVW":
        ws.column_dimensions[col_letter].width = 14
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_excel_with_tpt(res: dict, tpt_data: dict) -> BytesIO:
    """Export avec onglet Alloc + onglet TPT."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    # Crée d'abord le workbook avec l'onglet Alloc
    buf_alloc = export_excel(res)
    wb = openpyxl.load_workbook(buf_alloc)

    # Ajoute onglet TPT
    ws_tpt = wb.create_sheet("Analyse TPT")
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfill = PatternFill(fgColor="1F3864", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    dfont = Font(name="Calibri", size=11)

    CIC_LABELS = {
        "1": "Obligations souveraines", "2": "Obligations d'entreprises",
        "3": "Actions", "4": "Fonds d'investissement", "5": "Titres structurés",
        "7": "Trésorerie et dépôts", "8": "Prêts et hypothèques",
        "9": "Immobilisations corporelles", "A": "Futures", "B": "Options call",
        "D": "Swaps", "E": "Forwards", "TOTAL": "TOTAL",
    }
    tpt_headers = ["CIC", "Catégorie", "Val. marché (€)", "Duration", "SCR total (€)",
                   "SCR taux (€)", "SCR spread (€)", "SCR actions (€)", "SCR immo (€)", "SCR forex (€)",
                   "SCR actions cotées (€)", "SCR actions non cotées (€)",
                   "SCR/Actifs", "SCR taux %", "SCR spread %", "SCR actions %", "SCR immo %", "SCR forex %"]
    for col, h in enumerate(tpt_headers, 1):
        c = ws_tpt.cell(row=1, column=col, value=h)
        c.fill = hfill;
        c.font = hfont;
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")

    def e(v):
        return round(v * 1e6, 2) if v else 0

    def pct(v):
        return round(v * 100, 3) if v else 0

    row_i = 2
    for cic, data in tpt_data.items():
        if data.get("FA", 0) == 0 and cic != "TOTAL": continue
        vals = [cic, CIC_LABELS.get(cic, cic), e(data.get("FA", 0)),
                round(data.get("FB", 0), 2) if data.get("FB") else None,
                e(data.get("FC", 0)), e(data.get("FD", 0)), e(data.get("FE", 0)),
                e(data.get("FF", 0)), e(data.get("FG", 0)), e(data.get("FH", 0)),
                e(data.get("FI", 0)), e(data.get("FJ", 0)),
                pct(data.get("FK", 0)), pct(data.get("FL", 0)), pct(data.get("FM", 0)),
                pct(data.get("FN", 0)), pct(data.get("FO", 0)), pct(data.get("FP", 0))]
        fmts_tpt = [None, None, "#,##0.00", "0.00", "#,##0.00", "#,##0.00", "#,##0.00",
                    "#,##0.00", "#,##0.00", "#,##0.00", "#,##0.00", "#,##0.00",
                    "0.000%", "0.000%", "0.000%", "0.000%", "0.000%", "0.000%"]
        for col, (val, fmt) in enumerate(zip(vals, fmts_tpt), 1):
            c = ws_tpt.cell(row=row_i, column=col, value=val)
            c.font = dfont;
            c.border = border
            c.alignment = Alignment(horizontal="right" if col > 2 else "left", vertical="center")
            if fmt and val is not None: c.number_format = fmt
        row_i += 1

    ws_tpt.column_dimensions["A"].width = 8
    ws_tpt.column_dimensions["B"].width = 28
    for col_letter in "CDEFGHIJKLMNOPQR":
        ws_tpt.column_dimensions[col_letter].width = 16

    buf2 = BytesIO()
    wb.save(buf2)
    buf2.seek(0)
    return buf2


# ══════════════════════════════════════════════════════════════════
#  UI
# ══════════════════════════════════════════════════════════════════
st.markdown("""
<h1 style='color:#1F3864;font-family:Calibri,sans-serif;'>Suivi d'Allocation</h1>
<p style='color:#666;font-size:14px;'>Dépose le fichier Excel → tous les calculs automatiquement.</p>
<hr style='border:1px solid #ddd;'>
""", unsafe_allow_html=True)

uploaded = st.file_uploader("Fichier Excel (.xlsx)", type=["xlsx"])

if uploaded:
    try:
        file_bytes = uploaded.read()
        wb_chk = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True)
        sheet_names = wb_chk.sheetnames
        wb_chk.close()

        missing = [s for s in ["Portefeuille", "Retraitements", "Nantissement"] if s not in sheet_names]
        if missing:
            st.error(f"Onglets manquants : `{'`, `'.join(missing)}`")
            st.stop()
        if "KNL" not in sheet_names:
            st.warning("Onglet KNL absent — colonnes Engagements/Projection à 0.")
        if "TPT" not in sheet_names:
            st.warning("Onglet TPT absent — duration et SCR à 0.")

        with st.spinner("Calcul en cours..."):
            res, calc_warnings = compute(file_bytes)
        if calc_warnings:
            for w in calc_warnings:
                st.warning(w)
            # Recup tpt_scr pour affichage tableau TPT
            if "TPT" in sheet_names:
                wb_tpt = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
                ws_tpt_ui = wb_tpt["TPT"]
                rows_tpt_ui = []
                for row in ws_tpt_ui.iter_rows(min_row=2, values_only=True):
                    if not row or len(row) <= 153:
                        continue
                    rows_tpt_ui.append({
                        "Z": float(row[25]) if row[25] else 0.0,
                        "EX": str(row[153]).strip() if row[153] else "",
                        "EQ": float(row[146]) if row[146] else 0.0,
                        "CU": float(row[98]) if row[98] else 0.0,
                        "CZ": float(row[103]) if row[103] else 0.0,
                        "CW": float(row[100]) if row[100] else 0.0,
                        "CX": float(row[101]) if row[101] else 0.0,
                        "CY": float(row[102]) if row[102] else 0.0,
                        "DE": float(row[108]) if row[108] else 0.0,
                        "NAV": float(row[8]) if row[8] else 0.0,
                    })
                if rows_tpt_ui:
                    df_ui = pd.DataFrame(rows_tpt_ui)
                    nav_ui = df_ui["NAV"].iloc[0] if len(df_ui) > 0 else 1.0
                    corr_ui = np.array(
                        [[1, 0, 0, 0, 0.25], [0, 1, 0.75, 0.5, 0.25], [0, 0.75, 1, 0.75, 0.25], [0, 0.5, 0.75, 1, 0.25],
                         [0.25, 0.25, 0.25, 0.25, 1]])
                    scr_store = {}
                    for cic in ["1", "2", "3", "4", "5", "7", "8", "9", "A", "B", "D", "E"]:
                        sub = df_ui[df_ui["EX"] == cic]
                        z = sub["Z"].sum()
                        if z == 0: continue
                        fd = sub["CU"].sum() * nav_ui
                        fe = sub["CZ"].sum() * nav_ui
                        eq1 = sub["CW"].sum() * nav_ui
                        eq2 = sub["CX"].sum() * nav_ui
                        fg = sub["CY"].sum() * nav_ui
                        fh = sub["DE"].sum() * nav_ui
                        ff = float(np.sqrt(eq1 ** 2 + eq2 ** 2 + 0.75 * eq1 * eq2)) if (eq1 or eq2) else 0.0
                        v = np.array([fd, fe, ff, fg, fh])
                        fc = float(np.sqrt(v @ corr_ui @ v)) if v.any() else 0.0
                        dur = (sub["EQ"] * sub["Z"]).sum() / z
                        fb = (fd / z / 0.01) if z else None  # FB = FD/FA/1%
                        scr_store[cic] = {
                            "FA": z / 1e6,
                            "FB": fb,
                            "FC": fc / 1e6,
                            "FD": fd / 1e6,
                            "FE": fe / 1e6,
                            "FF": ff / 1e6,
                            "FG": fg / 1e6,
                            "FH": fh / 1e6,
                            "FI": eq1 / 1e6,  # SCR actions cotées
                            "FJ": eq2 / 1e6,  # SCR actions non cotées
                            "FK": fc / z if z else 0.0,
                            "FL": fd / z if z else 0.0,  # SCR taux %
                            "FM": fe / z if z else 0.0,  # SCR spread %
                            "FN": ff / z if z else 0.0,  # SCR actions %
                            "FO": fg / z if z else 0.0,  # SCR immo %
                            "FP": fh / z if z else 0.0,  # SCR forex %
                        }
                    # Total
                    z_t = df_ui["Z"].sum()
                    fd_t = df_ui["CU"].sum() * nav_ui;
                    fe_t = df_ui["CZ"].sum() * nav_ui
                    eq1_t = df_ui["CW"].sum() * nav_ui;
                    eq2_t = df_ui["CX"].sum() * nav_ui
                    fg_t = df_ui["CY"].sum() * nav_ui;
                    fh_t = df_ui["DE"].sum() * nav_ui
                    ff_t = float(np.sqrt(eq1_t ** 2 + eq2_t ** 2 + 0.75 * eq1_t * eq2_t)) if (eq1_t or eq2_t) else 0.0
                    v_t = np.array([fd_t, fe_t, ff_t, fg_t, fh_t])
                    fc_t = float(np.sqrt(v_t @ corr_ui @ v_t)) if v_t.any() else 0.0
                    fb_t = (fd_t / z_t / 0.01) if z_t else None
                    scr_store["TOTAL"] = {
                        "FA": z_t / 1e6,
                        "FB": fb_t,
                        "FC": fc_t / 1e6,
                        "FD": fd_t / 1e6,
                        "FE": fe_t / 1e6,
                        "FF": ff_t / 1e6,
                        "FG": fg_t / 1e6,
                        "FH": fh_t / 1e6,
                        "FI": eq1_t / 1e6,
                        "FJ": eq2_t / 1e6,
                        "FK": fc_t / z_t if z_t else 0.0,
                        "FL": fd_t / z_t if z_t else 0.0,
                        "FM": fe_t / z_t if z_t else 0.0,
                        "FN": ff_t / z_t if z_t else 0.0,
                        "FO": fg_t / z_t if z_t else 0.0,
                        "FP": fh_t / z_t if z_t else 0.0,
                    }
                    st.session_state["tpt_scr_data"] = scr_store

        # Controle Retraitements
        categories_connues = {
            normalize(B)
            for _, rtype, detail, B, _, _, _ in ROW_DEFS
            if rtype == "white" and detail == "normal" and B is not None
        }
        wb_chk2 = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        ws_rchk = wb_chk2["Retraitements"]
        classes_fichier = set()
        for row in ws_rchk.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 3 and row[2] is not None:
                val = str(row[2]).strip()
                if val.lower() not in ("classe d'actifs", "montant", ""):
                    classes_fichier.add(normalize(val))
        non_matches = classes_fichier - categories_connues
        if non_matches:
            st.warning(f"Classes Retraitements non reconnues : **{', '.join(sorted(non_matches))}**")

        g_tot = res[TOTAL_IDX]["G"]
        j_tot = res[TOTAL_IDX]["J"]
        p_tot = res[TOTAL_IDX].get("P") or 0.0
        # Duration moyenne S25 = SOMMEPROD(EQ;Z)/SOMME(Z) depuis TPT global
        dur_moy = res[TOTAL_IDX].get("S") or 0.0
        # Duration ALM = V25 = U25 + SOMME(V_i) = sum(U_i * P_i%)
        # P_i% = res[i]["Q"] = alloc projetee %
        dur_alm = sum(
            (res[i].get("U") or 0.0) * (res[i].get("Q") or 0.0)
            for i in range(len(ROW_DEFS)) if i in res and i != TOTAL_IDX
        )
        # SCR/Actifs W25 = FK21 total avec diversification
        scr_act = res[TOTAL_IDX].get("W") or 0.0
        # SCR taux = FD21/FA21 = scr_taux_total / fa_total
        scr_taux_kpi = (res[TOTAL_IDX].get("X") or 0.0) if False else 0.0  # placeholder

        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("Total Allocation", f"{g_tot:,.0f} M€")
        c2.metric("Alloc. projetée", f"{p_tot:,.0f} M€")
        c3.metric("Duration moyenne", f"{dur_moy:.3f}")
        c4.metric("Duration ALM", f"{dur_alm:.3f}")
        c5.metric("SCR/Actifs", f"{scr_act * 100:.2f}%")

        st.markdown("<br>", unsafe_allow_html=True)

        # Toggles colonnes
        col_t1, col_t2 = st.columns(2)
        show_ret = col_t1.checkbox("Retraitement", value=False)
        show_qy = col_t2.checkbox("Colonnes pilotage (Écart, Duration, SCR)", value=False)

        components.html(render_table(res, show_retraitement=show_ret, show_qy=show_qy),
                        height=800, scrolling=True)
        st.markdown("<br>", unsafe_allow_html=True)

        # Boutons export — juste sous le tableau
        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            st.download_button(
                label="📥 Exporter tableau Alloc",
                data=export_excel(res),
                file_name="suivi_allocation_calcule.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with col_dl2:
            tpt_data_dl = st.session_state.get("tpt_scr_data", {})
            st.download_button(
                label="📥 Exporter tableau + TPT",
                data=export_excel_with_tpt(res, tpt_data_dl),
                file_name="suivi_allocation_avec_tpt.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        # Camemberts côte à côte
        import plotly.graph_objects as go

        st.markdown("---")
        st.markdown("### Répartition par catégorie")
        colors_pie = LMG_COLORS

        g_tot_pie = res[TOTAL_IDX]["G"]
        p_tot_pie = res[TOTAL_IDX].get("P") or 1.0

        labels_a, vals_a, labels_p, vals_p, cols_a, cols_p = [], [], [], [], [], []
        for i, br in enumerate(BLUE_IDX):
            r = res.get(br)
            if not r: continue
            col = colors_pie[i % len(colors_pie)]
            h_pct = r.get("H") or 0.0  # alloc actuelle H%
            q_pct = r.get("Q") or 0.0  # alloc projetée Q% (déjà ajusté pour nanties)
            if h_pct > 0:
                labels_a.append(r["label"]);
                vals_a.append(round(h_pct * 100, 1));
                cols_a.append(col)
            if q_pct > 0:
                labels_p.append(r["label"]);
                vals_p.append(round(q_pct * 100, 1));
                cols_p.append(col)

        col_pie1, col_pie2 = st.columns(2)
        date_actuelle = "31/12/2025"
        date_projetee = "31/12/2029"
        for col_pie, labels, vals, cols, titre in [
            (col_pie1, labels_a, vals_a, cols_a, f"Allocation actuelle<br>{date_actuelle}"),
            (col_pie2, labels_p, vals_p, cols_p, f"Allocation projetée<br>{date_projetee}"),
        ]:
            # Affiche la valeur Q% originale comme label (même logique qu'Excel)
            text_labels = [f"{v:.1f}%" for v in vals]
            fig = go.Figure(go.Pie(
                labels=labels, values=vals, hole=0.3,
                marker=dict(colors=cols),
                text=text_labels,
                textinfo="text", textposition="inside",
                hovertemplate="%{label}<br>%{text}<extra></extra>",
            ))
            fig.update_layout(title=dict(text=titre, x=0.5, font=dict(size=13)),
                              showlegend=True, margin=dict(t=50, b=10, l=10, r=10), height=400)
            with col_pie:
                st.plotly_chart(fig, use_container_width=True)

        # ── Tableau TPT (SCR par catégorie CIC) ──────────────────
        st.markdown("---")
        st.markdown("### Analyse TPT — SCR par catégorie CIC")
        CIC_LABELS = {
            "1": "Obligations souveraines", "2": "Obligations d'entreprises",
            "3": "Actions", "4": "Fonds d'investissement",
            "5": "Titres structurés", "7": "Trésorerie et dépôts",
            "8": "Prêts et hypothèques", "9": "Immobilisations corporelles",
            "A": "Contrats à terme (futures)", "B": "Options d'achat (call)",
            "D": "Contrats d'échange (swaps)", "E": "Contrats à terme (forwards)",
            "TOTAL": "TOTAL",
        }
        if "tpt_scr_data" in st.session_state:
            tpt_table = []
            for cic, data in st.session_state["tpt_scr_data"].items():
                if data.get("FA", 0) == 0 and cic != "TOTAL":
                    continue


                def e(v):
                    return round(v * 1e6, 2) if v else 0  # M€ → €


                def pct(v):
                    return f"{v * 100:.3f}%" if v else "0.000%"


                tpt_table.append({
                    "CIC": cic,
                    "Catégorie": CIC_LABELS.get(cic, cic),
                    "Val. marché (€)": e(data.get("FA", 0)),
                    "Duration": round(data.get("FB", 0), 2) if data.get("FB") else "-",
                    "SCR total (€)": e(data.get("FC", 0)),
                    "SCR taux (€)": e(data.get("FD", 0)),
                    "SCR spread (€)": e(data.get("FE", 0)),
                    "SCR actions (€)": e(data.get("FF", 0)),
                    "SCR immo (€)": e(data.get("FG", 0)),
                    "SCR forex (€)": e(data.get("FH", 0)),
                    "SCR actions cotées (€)": e(data.get("FI", 0)),
                    "SCR actions non cotées (€)": e(data.get("FJ", 0)),
                    "SCR/Actifs": pct(data.get("FK", 0)),
                    "SCR taux %": pct(data.get("FL", 0)),
                    "SCR spread %": pct(data.get("FM", 0)),
                    "SCR actions %": pct(data.get("FN", 0)),
                    "SCR immo %": pct(data.get("FO", 0)),
                    "SCR forex %": pct(data.get("FP", 0)),
                })
            st.dataframe(pd.DataFrame(tpt_table), use_container_width=True, hide_index=True)
        else:
            st.info("Onglet TPT requis pour ce tableau.")

    except KeyError as e:
        st.error(f"Onglet introuvable : {e}")
    except Exception as e:
        st.error(f"Erreur : {e}")
        with st.expander("Details"):
            st.exception(e)
else:
    st.info("En attente du fichier Excel...")
    with st.expander("Onglets requis"):
        st.markdown("""
| Onglet | Colonnes utilisées |
|---|---|
| **Portefeuille** | F, W, Y, AK |
| **Nantissement** | F, W, Y |
| **Retraitements** | A, B, C, D |
| **KNL** *(optionnel)* | B, K, M, N, E1 |
| **TPT** *(optionnel)* | I, Z, CT, EQ, CU, CW, CX, CY, CZ, DE, EX |
        """)